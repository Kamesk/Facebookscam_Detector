from collections import OrderedDict
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
import re
import openpyxl
from openpyxl import workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import TimeoutException
import time
from datetime import date, datetime, timedelta
from selenium.webdriver.chrome.service import Service
import traceback
import datetime
import os

# Function to get the row count in a specific sheet of an Excel file
def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.max_row

# Function to get the column count in a specific sheet of an Excel file
def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.max_column

# Function to read data from a specific cell in a sheet of an Excel file
def readData(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum, column=columnno).value

# Function to write data to a specific cell in a sheet of an Excel file
def writeData(file, sheetName, rownum, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)

# Set up Chrome WebDriver
service = Service()
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=options)

# Navigate to the Facebook Marketplace web page
url = "https://www.facebook.com/marketplace/london/search/?query=london%20rental%20property"
driver.get(url)

# Wait for the page to load
time.sleep(30)

# Function to scroll down the page to load more content
def scroll_to_bottom():
    last_height = driver.execute_script("return document.body.scrollHeight")
    while True:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        try:
            WebDriverWait(driver, 10).until(
                lambda driver: driver.execute_script("return document.body.scrollHeight") > last_height
            )
            last_height = driver.execute_script("return document.body.scrollHeight")
        except TimeoutException:
            # Break the loop if there is no more new content loaded
            break

# Scroll down the page to load more content
scroll_to_bottom()

# Define the list of classes for <img> and <span> elements
img_classes_to_match = "xt7dq6l xl1xv1r x6ikm8r x10wlt62 xh8yej3"
span_classes_to_match = "x193iq5w xeuugli x13faqbe x1vvkbs x1xmvt09 x1lliihq x1s928wv xhkezso x1gmr53x x1cpjm7i x1fgarty x1943h6x xudqn12 x676frb x1lkfr7t x1lbecb7 x1s688f xzsf02u"

# Find all <img> elements with the specified class
img_elements = driver.find_elements(By.XPATH, f'//img[contains(@class, "{img_classes_to_match}")]')

# Find all <span> elements with the specified class
span_elements = driver.find_elements(By.XPATH, f'//span[contains(@class, "{span_classes_to_match}")]')

# Lists to store the extracted values
titles = []
src_urls = []
prices = []

# Loop through each element and extract the title, src, and price
for i in range(len(img_elements)):
    # Extract the title from the <span> element
    title = span_elements[i].text
    
    # Extract the src from the <img> element
    src_url = img_elements[i].get_attribute('src')
    
    # Extract the price from the alt_text using regular expressions
    alt_text = img_elements[i].get_attribute('alt')
    price_match = re.search(r'£(\d+)', alt_text)
    price = price_match.group(1) if price_match else None

    # Store the extracted values in lists
    titles.append(title)
    src_urls.append(src_url)
    prices.append(price)

# Close the Chrome WebDriver
driver.quit()

# Write the extracted values to the Excel file
path = "D:\\Westiminster\\Final project\\FB_SCAM\\Data.xlsx"  # Replace with the actual path of the Excel file
for i in range(len(titles)):
    writeData(path, "Sheet1", i+1, 1, titles[i])  # Write the title to column 1
    writeData(path, "Sheet1", i+1, 2, src_urls[i])  # Write the src_url to column 2
    writeData(path, "Sheet1", i+1, 3, prices[i])  # Write the price to column 3
